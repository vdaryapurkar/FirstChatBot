"""Build the downloadable output workbook: deterministic comparison/triage
tables (computed in xlsx_ingest.py) plus Claude's root-cause narrative
(from claude_client.py), formatted as a professional report.
"""

from __future__ import annotations

from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.xlsx_ingest import (
    CATEGORY_BUY_SELL,
    CATEGORY_OTHER,
    MISMATCH_PRIMARY_COLUMN_PRIORITY,
    STRUCTURAL_MISMATCHTYPES,
    ordered_mismatchtypes,
)

FONT_NAME = "Arial"
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
GROUP_FILL = PatternFill("solid", fgColor="D9D9D9")

# One fill per mismatchtype so the different issue groups within a sheet are
# visually distinguishable, not just labeled. Unrecognized mismatchtype
# values fall back to the "mismatch" fill.
MISMATCHTYPE_FILLS = {
    "mismatch": PatternFill("solid", fgColor="FCE4E4"),           # red-ish: genuine value diff
    "new_post": PatternFill("solid", fgColor="DCE6F7"),           # blue-ish: new position in post
    "missing_position_post": PatternFill("solid", fgColor="FDEBD3"),  # amber-ish: dropped position
}
DEFAULT_MISMATCHTYPE_FILL = MISMATCHTYPE_FILLS["mismatch"]
MISMATCHTYPE_LABELS = {
    "mismatch": "Mismatch",
    "new_post": "New position (post only)",
    "missing_position_post": "Missing position (post)",
}
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True)
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, italic=True, color="595959")
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_BODY_FONT = Font(name=FONT_NAME, size=10, bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)


def _style_header_row(ws, row: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autosize(ws, widths: list[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_report(triage: dict[str, dict], grouped: dict[str, dict], analysis: dict,
                  sources: list[str], output_path: str,
                  process_types_by_file: dict[str, str] | None = None,
                  credit_subtypes_by_file: dict[str, str] | None = None):
    """triage: the FULL, uncapped per-column output of xlsx_ingest.compute_triage
    (every row, not the token-budget-limited sample that was sent to Claude).
    grouped: the FULL, uncapped {mismatchtype: {group_key: [rows]}} output of
    compute_triage -- every flagged row pooled/deduped across the columns it
    triggered, one entry per distinct issue (see compute_triage's docstring
    for what group_key means per mismatchtype).
    analysis: the parsed tool-call dict from claude_client.run_analysis (the
    'result' key: summary/triage_categories/root_causes/recommendations).
    sources: list of original uploaded file names included in this report.
    process_types_by_file: {source_file: "Valuation"|"Settlement"|"NetValuation"|"Credit"|"Unknown"},
    from xlsx_ingest.build_data_digest()'s "process_types_by_file".
    credit_subtypes_by_file: {source_file: "CreditExposure"|"CreditExposureInterface"|
    "CreditForwardAnalysis"|"CreditForwardModel"}, from build_data_digest()'s
    "credit_subtypes_by_file" -- drives the RVA sheet (skipped entirely if empty).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # A column with zero TRUE rows reconciled cleanly (e.g. exposureqty_mismatch,
    # expprice_mismatch on a run where those never differ) -- no issue to
    # report, so it's left out of the "clean" count message below rather than
    # padding the workbook with tabs that only ever show "No issue".
    clean_column_count = sum(1 for issue in triage.values() if not issue["true_rows"])

    _build_summary_sheet(wb, grouped, analysis, sources, process_types_by_file or {}, clean_column_count)
    _build_root_cause_sheet(wb, analysis, grouped)
    for mismatchtype in ordered_mismatchtypes(grouped):
        _build_mismatchtype_sheet(wb, mismatchtype, grouped[mismatchtype])
    _build_rva_sheet(wb, grouped, credit_subtypes_by_file or {}, analysis)

    wb.save(output_path)


def _ordered_group_keys(mismatchtype: str, groups: dict[str, list]) -> list[str]:
    """For structural mismatchtypes, group_key is a position-type category:
    BUY/SELL first, then Other, then anything else found. For everything
    else (usually "mismatch"), group_key is a column name: the configured
    "primary driver" columns first (see MISMATCH_PRIMARY_COLUMN_PRIORITY),
    then the rest alphabetically."""
    present = set(groups.keys())
    if mismatchtype in STRUCTURAL_MISMATCHTYPES:
        ordered = [c for c in (CATEGORY_BUY_SELL, CATEGORY_OTHER) if c in present]
    else:
        ordered = [c for c in MISMATCH_PRIMARY_COLUMN_PRIORITY if c in present]
    ordered += sorted(present - set(ordered))
    return ordered


def _build_summary_sheet(wb, grouped, analysis, sources, process_types_by_file, clean_column_count=0):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Recon-CI Triage Analysis - Summary"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  |  Sources: {', '.join(sources)}"
    ws["A2"].font = SUBTITLE_FONT

    r = 4
    ws.cell(row=r, column=1, value="Source file")
    ws.cell(row=r, column=2, value="Process type")
    _style_header_row(ws, r, 2)
    r += 1
    for src in sources:
        ws.cell(row=r, column=1, value=src).font = BODY_FONT
        ws.cell(row=r, column=2, value=process_types_by_file.get(src, "Unknown")).font = BODY_FONT
        for j in (1, 2):
            ws.cell(row=r, column=j).border = BORDER
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Executive Summary").font = BOLD_BODY_FONT
    r += 1
    summary_start = r
    summary_end = summary_start + 4
    ws.merge_cells(start_row=summary_start, start_column=1, end_row=summary_end, end_column=6)
    cell = ws.cell(row=summary_start, column=1)
    cell.value = analysis.get("summary", "")
    cell.font = BODY_FONT
    cell.alignment = WRAP
    r = summary_end + 1

    # One synopsis + description per distinct issue (see core/llm_schema.py's
    # "bug_reports"), each meant to be pasted into its OWN separate TFS/Azure
    # DevOps ticket -- not merged into a single combined bug report.
    bug_reports = [b for b in analysis.get("bug_reports", []) if b.get("synopsis") or b.get("description")]
    if bug_reports:
        r += 1
        ws.cell(row=r, column=1,
                value="Bug Reports (copy each into TFS to file a separate investigation ticket)").font = BOLD_BODY_FONT
        r += 1
        for i, bug in enumerate(bug_reports, start=1):
            ws.cell(row=r, column=1, value=f"Bug {i} - Synopsis").font = BOLD_BODY_FONT
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
            c = ws.cell(row=r, column=1, value=bug.get("synopsis", ""))
            c.font = BODY_FONT
            c.alignment = WRAP
            r += 1
            ws.cell(row=r, column=1, value=f"Bug {i} - Description").font = BOLD_BODY_FONT
            r += 1
            desc_start = r
            desc_end = desc_start + 9
            ws.merge_cells(start_row=desc_start, start_column=1, end_row=desc_end, end_column=6)
            c = ws.cell(row=desc_start, column=1, value=bug.get("description", ""))
            c.font = BODY_FONT
            c.alignment = WRAP
            r = desc_end + 2

    if clean_column_count:
        r += 1
        plural = "s" if clean_column_count != 1 else ""
        ws.cell(row=r, column=1,
                value=(f"{clean_column_count} other mismatch column{plural} reconciled cleanly "
                       "(no differences found) and are omitted from this report.")).font = SUBTITLE_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        r += 1

    _autosize(ws, [24, 26, 12, 10, 24, 60])
    ws.row_dimensions[1].height = 22


def _count_lookup(grouped: dict[str, dict]) -> dict[tuple, int]:
    """{(mismatchtype, group_key): row count} across every issue -- used to
    add a Count column to sheets driven by the LLM's analysis (which doesn't
    itself carry a row count) without duplicating the full issue table the
    Summary sheet used to show (now dropped in favor of Root Cause Analysis,
    which already lists every issue)."""
    return {
        (mismatchtype, group_key): len(rows)
        for mismatchtype, groups in grouped.items()
        for group_key, rows in groups.items()
    }


def _build_root_cause_sheet(wb, analysis, grouped):
    ws = wb.create_sheet("Root Cause Analysis")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Root Cause Analysis"
    ws["A1"].font = TITLE_FONT

    headers = ["Issue", "Column", "Mismatch Type", "Category", "Count", "Process Type", "Explanation", "Evidence", "Affected Scope", "Confidence"]
    header_row = 3
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    _style_header_row(ws, header_row, len(headers))

    counts = _count_lookup(grouped)
    r = header_row + 1
    for rc in analysis.get("root_causes", []):
        evidence = "\n".join(f"- {e}" for e in rc.get("evidence", []))
        mismatchtype = rc.get("mismatchtype", "")
        group_key = rc.get("category") or rc.get("column")
        count = counts.get((mismatchtype, group_key), "")
        row_vals = [
            rc.get("issue", ""),
            rc.get("column", ""),
            MISMATCHTYPE_LABELS.get(mismatchtype, mismatchtype),
            rc.get("category", "") or "-",
            count,
            rc.get("process_type", ""),
            rc.get("explanation", ""),
            evidence,
            rc.get("affected_scope", ""),
            rc.get("confidence", ""),
        ]
        for j, v in enumerate(row_vals, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = WRAP
        ws.row_dimensions[r].height = 90
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Recommendations").font = BOLD_BODY_FONT
    r += 1
    for rec in analysis.get("recommendations", []):
        ws.cell(row=r, column=1, value=f"- {rec}").font = BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
        ws.cell(row=r, column=1).alignment = WRAP
        r += 1

    _autosize(ws, [28, 16, 22, 12, 10, 16, 45, 40, 22, 12])


def _build_mismatchtype_sheet(wb, mismatchtype: str, groups: dict[str, list]):
    """One sheet per mismatchtype ("Mismatch", "New position (post only)",
    "Missing position (post)", ...), with rows grouped by issue -- a
    position-type category for structural mismatchtypes, the primary
    driving column for everything else -- instead of one tab per literal
    mismatch column. A row that triggered several columns for the same
    underlying reason appears once, under its group, with every triggered
    column listed (see compute_triage())."""
    is_structural = mismatchtype in STRUCTURAL_MISMATCHTYPES
    label = MISMATCHTYPE_LABELS.get(mismatchtype, mismatchtype)
    ws = wb.create_sheet(label[:31])
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    ws["A1"] = f"Triage: {label}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Each position appears once, pooled across every mismatch column it "
                "triggered, grouped by position type." if is_structural else
                "Each position appears once, pooled across every mismatch column that "
                "moved together for the same reason, grouped by the driving column.")
    ws["A2"].font = SUBTITLE_FONT

    all_rows = [row for rows in groups.values() for row in rows]
    id_cols: list[str] = []
    seen = set()
    for row in all_rows:
        for k in row.keys():
            if k in ("_source_file", "_sheet_name", "_process_type", "_mismatchtype",
                      "_positiontype_category", "_triggered_columns", "_differences", "_difference"):
                continue
            if k.lower().endswith("_mismatch"):
                continue
            if k not in seen:
                seen.add(k)
                id_cols.append(k)

    diff_cols = sorted({c for row in all_rows for c, v in row.get("_differences", {}).items() if v is not None})

    headers = ["source_file", "sheet_name", "process_type", "triggered_columns"] + id_cols + [f"diff: {c}" for c in diff_cols]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    _style_header_row(ws, header_row, len(headers))

    fill = MISMATCHTYPE_FILLS.get(mismatchtype, DEFAULT_MISMATCHTYPE_FILL)
    r = header_row + 1
    for group_key in _ordered_group_keys(mismatchtype, groups):
        rows = groups[group_key]
        if not rows:
            continue
        if is_structural:
            group_label = f"CATEGORY - {group_key}"
        else:
            triggered_union = sorted({c for row in rows for c in row.get("_triggered_columns", [])})
            others = [c for c in triggered_union if c != group_key]
            group_label = f"ISSUE - {group_key}" + (f"  (also triggers: {', '.join(others)})" if others else "")
        ws.cell(row=r, column=1, value=f"{group_label}  ({len(rows)} rows)").font = BOLD_BODY_FONT
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
        ws.cell(row=r, column=1).fill = GROUP_FILL
        r += 1
        for row in rows:
            values = [
                row.get("_source_file", ""),
                row.get("_sheet_name", ""),
                row.get("_process_type", "Unknown"),
                ", ".join(row.get("_triggered_columns", [])),
            ]
            values += [row.get(k, "") for k in id_cols]
            diffs = row.get("_differences", {})
            values += [diffs.get(c) for c in diff_cols]
            for j, v in enumerate(values, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = BODY_FONT
                c.border = BORDER
                c.fill = fill
            r += 1

    _autosize(ws, [16, 14, 14, 30] + [14] * len(id_cols) + [14] * len(diff_cols))


# Display order for the RVA sheet's four Credit sub-processes -- see
# xlsx_ingest.CREDIT_SUBTYPES_ORDERED.
CREDIT_SUBTYPE_ORDER = ["CreditExposure", "CreditExposureInterface", "CreditForwardAnalysis", "CreditForwardModel"]


def _build_rva_sheet(wb, grouped: dict[str, dict], credit_subtypes_by_file: dict[str, str], analysis: dict):
    """RVA: a rollup consolidating findings across the four Credit
    sub-process files (creditexposure, creditexposureinterface,
    creditforwardanalysis, creditforwardmodel -- each its own uploaded file,
    identified by xlsx_ingest.detect_credit_subtype()). Skipped entirely
    when no Credit files were uploaded this run. A sub-process with no
    flagged rows still gets a row here noting it reconciled cleanly, rather
    than being silently omitted -- so the rollup always accounts for every
    Credit file that was actually uploaded."""
    if not credit_subtypes_by_file:
        return

    ws = wb.create_sheet("RVA")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "RVA - Credit Reconciliation Rollup"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Findings consolidated across the Credit sub-process files: CreditExposure, "
                "CreditExposureInterface, CreditForwardAnalysis, CreditForwardModel.")
    ws["A2"].font = SUBTITLE_FONT

    # Every pooled row for a Credit file carries its sub-process on
    # "_credit_subtype" (see compute_triage()); regroup by that here.
    by_subtype: dict[str, dict[tuple, list]] = {}
    for mismatchtype, groups in grouped.items():
        for group_key, rows in groups.items():
            for row in rows:
                subtype = row.get("_credit_subtype")
                if not subtype:
                    continue
                by_subtype.setdefault(subtype, {}).setdefault((mismatchtype, group_key), []).append(row)

    file_by_subtype = {v: k for k, v in credit_subtypes_by_file.items()}
    present = set(credit_subtypes_by_file.values()) | set(by_subtype.keys())
    ordered_subtypes = [s for s in CREDIT_SUBTYPE_ORDER if s in present]
    ordered_subtypes += sorted(present - set(ordered_subtypes))

    cat_by_key = {
        (c.get("column"), c.get("mismatchtype")): c for c in analysis.get("triage_categories", [])
        if c.get("column")
    }
    struct_cat_by_key = {
        (c.get("mismatchtype"), c.get("category")): c for c in analysis.get("triage_categories", [])
        if c.get("category") and not c.get("column")
    }

    headers = ["Sub-process", "Source file", "Issue (column(s))", "Mismatch Type", "Category", "Count", "Description"]
    header_row = 4
    for j, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=j, value=h)
    _style_header_row(ws, header_row, len(headers))

    r = header_row + 1
    for subtype in ordered_subtypes:
        source_file = file_by_subtype.get(subtype, "")
        issues = by_subtype.get(subtype, {})
        if not issues:
            row_vals = [subtype, source_file, "-", "-", "-", 0, "No differences observed between pre and post."]
            for j, v in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = BODY_FONT
                c.border = BORDER
                c.alignment = WRAP
            r += 1
            continue

        for mismatchtype, group_key in sorted(issues.keys()):
            rows = issues[(mismatchtype, group_key)]
            is_structural = mismatchtype in STRUCTURAL_MISMATCHTYPES
            triggered_columns = sorted({c for row in rows for c in row.get("_triggered_columns", [])})
            if is_structural:
                cat = struct_cat_by_key.get((mismatchtype, group_key), {})
                category_display = group_key
            else:
                cat = cat_by_key.get((group_key, mismatchtype), {})
                category_display = "-"
            row_vals = [
                subtype,
                source_file,
                ", ".join(triggered_columns),
                MISMATCHTYPE_LABELS.get(mismatchtype, mismatchtype),
                category_display,
                len(rows),
                cat.get("description", ""),
            ]
            for j, v in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = BODY_FONT
                c.border = BORDER
                c.alignment = WRAP
            r += 1

    _autosize(ws, [22, 34, 26, 20, 12, 10, 50])
    ws.row_dimensions[1].height = 22
